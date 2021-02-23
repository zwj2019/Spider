import os
import logging
import traceback
from urllib.parse import quote
import yagmail
from openpyxl import load_workbook

import EBCodeSearch.settings as settings

logger = logging.getLogger('spiderLogger')


def load_key_words():
    f = open('extra_settings.txt', 'r', encoding='utf-8')
    line = f.readline() # 关键字
    while line:
        key_word_list = line.split(',')
        line = f.readline().strip() # 关键字描述
        for key_word in key_word_list:
            key_word = key_word.strip()
            if key_word is not None and len(key_word) > 0:
                settings.KEY_WORDS_DETAIL[key_word] = line

        line = f.readline()
    f.close()


def prepare_key_words(spider_name: str):
    # 对参数进行URL编码
    kw_list = [quote(kw) for kw in settings.KEY_WORDS]

    if str.startswith(spider_name, "gitee"):
        return " ".join(kw_list)

    return "+".join(kw_list)


def send_email():
    try:
        yag = yagmail.SMTP(user=settings.EMAIL_USER, password=settings.EMAIL_PASSWORD,
                           host=settings.EMAIL_SERVER)

        attachments = []

        if os.path.exists('config.txt'):
            attachments.append('config.txt')
        data_root = settings.OUTPUT_DIR

        for file in os.listdir(data_root):
            file_path = os.path.join(data_root, file)
            if os.path.isfile(file_path):
                attachments.append(file_path)

        # 构建关键字信息用于发送邮件
        contents = "本次各平台查询情况为：\n"
        contents += "Github : %d条\n" % settings.GITHUB_COUNT
        contents += "Gitee : %d条\n" % settings.GITEE_COUNT
        contents += "Gitlab : %d条\n" % settings.GITLAB_COUNT

        contents += "\n字段信息：\n" + str(settings.RESULT_DETAIL)
        yag.send(to=settings.TO, subject='Python Email Test', contents=contents, attachments=attachments)
        logger.info("Send Email to " + str(settings.TO))

    except Exception as e:
        logger.warning("Fail to send email!")
        print(traceback.format_exc())



def read_excel(file_path):

    wb = load_workbook(file_path)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[-1])
    assert wb.max_row > 1
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            print(cell. value, end=' ')
        print()
